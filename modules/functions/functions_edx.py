"""
Functions used in EDX interactive plot using dash module to detach it completely from Jupyter Notebooks.
Internal use for Institut Néel and within the MaMMoS project, to export and read big datasets produced at Institut Néel.

@Author: William Rigaut - Institut Néel (william.rigaut@neel.cnrs.fr)
"""

from pathlib import Path
import numpy as np
import plotly.graph_objects as go
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

from ..functions.functions_shared import *

def make_path_name(
    foldername, x_pos, y_pos, start_x=-40, start_y=-40, step_x=5, step_y=5
):
    """Converting X and Y real coordinates to the indices used by bruker to label files during edx scans

    Parameters
    ----------
    foldername : STR
        Folder that contains all the EDX datafiles in the .spx format.
        Used to read data in the correct path
    x_pos, ypos : INT, INT
        Horizontal position X (in mm) and vertical position Y (in mm) on the sample.
        The EDX scan saved the datafiles labeled by two numbers (a, b) corresponding to the scan number in the x and y positions.
    start_x, start_y : INT, INT optional
        Starting X and Y positions of the EDX scan, by default it will always start at (-40, -40)
    step_x, step_y : INT, INT optional
        X and Y steps for the EDX scan, by default it will be setup to 5 and 5 (mm)

    Returns
    -------
    path_name : STR
        Full path of the spx file at the (X, Y) position
    """
    # Converting wafer position to datafile label number
    x_idx, y_idx = int((x_pos - start_x) / step_x + 1), int(
        (y_pos - start_y) / step_y + 1
    )

    path_name = Path(f"{foldername}/Spectrum_({x_idx},{y_idx}).spx")
    if path_name.is_file():
        return path_name
    else:
        path_name = Path(f"{foldername}/Spectre_({x_idx},{y_idx}).spx")
        if path_name.is_file():
            return path_name
        else:
            raise FileNotFoundError("Couldn't find spectrum file, check your folder")


def create_result_list(element_list, results_ext):
    result_list = []

    for i, elm in enumerate(results_ext):
        if int(elm[0][1]) in [nb[0] for nb in element_list]:
            element_name = element_list[i][1]
        for attrb in elm[1:]:
            result_list.append([element_name, round(float(attrb[1]) * 100, 2)])
    return result_list


def get_spectra_spx(spx_file):
    """Reading spx datafile exported by the BRUKER software. The function will iter through the file,
    fetching the needed attributs in the metadata and the actual data to create a Scatter Figure with the plotly module

    Parameters
    ----------
    spx_file : STR
        Path containing the spx file

    Returns
    -------
    edx_spectra : NP.ARRAY
        Numpy array containing the counts as a function of energy
    """
    # Metadata to keep for displaying quantification results
    energy_step = 0.0
    zero_energy = 0.0
    results_ext = []
    element_list = []

    # Iteration through the XML datafile, searching for important metadata and EDX spectra
    tree = ET.parse(spx_file)
    root = tree.getroot()
    for elm in root.iter():
        if elm.tag == "PrimaryEnergy":
            voltage = int(float(elm.text))
        elif elm.tag == "WorkingDistance":
            working_distance = float(elm.text)
        elif elm.tag == "CalibLin":
            energy_step = float(elm.text)
        elif elm.tag == "CalibAbs":
            zero_energy = float(elm.text)
        elif elm.tag == "Channels":
            edx_spectra = np.array(
                [
                    ((i + 1) * energy_step + zero_energy, int(counts))
                    for i, counts in enumerate(elm.text.split(","))
                ]
            )
        # Getting element quantification results
        elif elm.tag == "ClassInstance" and elm.get("Name") == "Results":
            for child in elm.iter():
                if child.tag == "Result":
                    results_ext.append([])
                elif child.tag in ["Atom", "AtomPercent"]:
                    if child.tag == "Atom" and int(child.text) < 10:
                        results_ext[-1].append((child.tag, f"0{child.text}"))
                    else:
                        results_ext[-1].append((child.tag, child.text))
                elif child.tag == "ExtResults":
                    break
        elif elm.tag == "ClassInstance" and elm.get("Name") == "Elements":
            for child in elm.iter():
                if child.get("Type") == "TRTPSEElement":
                    name_elm = child.get("Name")
                    for nb in child.iter():
                        if nb.tag == "Element":
                            nb_elm = nb.text
                    element_list.append([int(nb_elm), name_elm])

    result_list = create_result_list(element_list, results_ext)

    metadata_lst = [voltage, working_distance, energy_step, zero_energy, result_list]

    return edx_spectra, metadata_lst


def make_heatmap(data, element_edx, start_x=-40, start_y=-40, step_x=5, step_y=5):
    """Fetching every X and Y positions in the xlsx datafile for a given element

    Parameters
    ----------
    data : LIST
        List containing every line of the xlsx file. For each line the name scan is followed by the quantification results for each element
    element_edx : STR
        Element that will be displayed in the heatmap
    start_x, start_y : INT, INT optional
        Starting X and Y positions of the EDX scan, by default it will always start at (-40, -40)
    step_x, step_y : INT, INT optional
        X and Y steps for the EDX scan, by default it will be setup to 5 and 5 (mm)

    Returns
    -------
    X_POS, Y_POS : LIST
        List containing every (X, Y) scan positions for the EDX map
    ELM : LIST
        Values of the quantification result at each position
    """
    X_POS = []
    Y_POS = []
    ELM = []

    for row in data:
        if row[0] == "Spectrum":
            index = row.index(element_edx)
        elif row[0].startswith("Spectrum_"):
            x_index, y_index = row[0].split("(")[-1].split(")")[0].split(",")
            x_pos, y_pos = (int(x_index) - 1) * step_x + start_x, (
                int(y_index) - 1
            ) * step_y + start_y
            if np.abs(x_pos) + np.abs(y_pos) <= 60:
                X_POS.append(x_pos)
                Y_POS.append(y_pos)
                ELM.append(float(row[index]))

    return X_POS, Y_POS, ELM


def generate_spectra(foldername, x_pos, y_pos):
    """Reading the EDX data from .xml datafile exported by the BRUKER software. The function will iter through the file,
    fetching the needed attributs in the metadata and the actual data to create a Scatter Figure with the plotly module

    Parameters
    ----------
    foldername : STR
        Folder that contains all the EDX datafiles in the .xml format.
        Used to read data in the correct path
    x_pos, ypos : INT, INT
        Horizontal position X (in mm) and vertical position Y (in mm) on the sample.
        The EDX scan saved the datafiles labeled by two numbers (a, b) corresponding to the scan number in the x and y positions.

    Returns
    -------
    fig : FIGURE OBJ
        Figure object from plotly.graph_objects containing a Scatter plot
    """
    # Defining a empty Figure object to send when certain conditions are not met
    empty_fig = go.Figure(data=go.Scatter())
    empty_fig.update_layout(height=750, width=1500)
    empty_meta = go.layout.Annotation()

    # If the user did not select a data folder, the displayed graph will be empty
    if foldername is None:
        return empty_fig, empty_meta

    # getting the spectrum from spx file
    spx_file = make_path_name(foldername, x_pos, y_pos)
    edx_spectra, metadata = get_spectra_spx(spx_file)

    # Creating the scatter plot with plotly.graph_objects library
    fig = go.Figure(
        data=[
            go.Scatter(
                x=[elm[0] for elm in edx_spectra],
                y=[elm[1] for elm in edx_spectra],
                marker_color="red",
            )
        ]
    )

    # Creating the metadata annotation within the plot
    results_str = "".join(
        [str(elm[0]) + ": " + str(elm[1]) + " at.%<br>" for elm in metadata[4]]
    )
    annoted = "Voltage: {:} keV<br>Working Dist.: {:.5f} mm<br>Zero Energy: {:.3f} keV<br>{:}".format(
        metadata[0], metadata[1], metadata[3], results_str
    )

    meta = go.layout.Annotation(
        text=annoted,
        align="left",
        showarrow=False,
        xref="paper",
        yref="paper",
        x=0.95,
        y=0.95,
        bordercolor="black",
        borderwidth=1,
    )

    return fig, meta


def get_elements(foldername, with_plot=False):
    """Reading from the Global Spectrum Results.xlsx file, the function will return all the element that were used for the quantification
    but excluding the element that were deconvoluted

    Parameters
    ----------
    foldername : STR
        Folder that contains all the EDX datafiles in the .xml format.
        Used to read data in the correct path
    with_plot : BOOL (optional)
        False by default, if True, the function will also return another LIST containing the data after reading the .xlsx file

    Returns
    -------
    elm_options : LIST
        List of element that were used for the quantification
    LIST_DATA : LIST (optional)
        List containing the data read inside the .xlsx file
    """

    # Reading in .xlsx file using openpyxl library, checking if file exists, if not returns empty list
    try:
        wb = load_workbook(filename= Path(f"{foldername}/Global spectrum results.xlsx"))
    except FileNotFoundError:
        try:
            wb = load_workbook(filename= Path(f"{foldername}/Résultats globaux.xlsx"))
        except FileNotFoundError:
                return []
    ws = wb.active

    # putting all the data inside lists to obtain a list of each line
    LIST_DATA = []
    for i, row in enumerate(ws.values):
        LIST_DATA.append(row)

    # checking all the element that have been quantified by excluding the deconvoluted ones
    index_tab = []
    for index, elm in enumerate(LIST_DATA[-3]):
        if index > 0 and elm is not None:
            if float(elm) != 0.0:
                index_tab.append(index)
    elm_options = [LIST_DATA[0][index] for index in index_tab]

    if with_plot:
        return elm_options, LIST_DATA
    else:
        return elm_options


def generate_heatmap(folderpath_edx, element_edx, z_min=None, z_max=None):
    """Plotting results from element quantification inside the Global Spectrum Results.xlsx file

    Parameters
    ----------
    folderpath_edx : STR
        Folder that contains the Global Spectrum Results.xlsx file to read quantification results
        Used to read data in the correct path
    element_edx : STR
        Element to display the concentration in at% from the xlsx file as a function of (X, Y) positions

    Returns
    -------
    fig : FIGURE OBJ
        Figure object from plotly.graph_objects containing a Heatmap plot
    """
    # Defining a dummy Figure object to send when certains conditions are not met
    empty_fig = go.Figure(data=go.Heatmap())
    empty_fig.update_layout(height=750, width=750)
    if folderpath_edx is None or element_edx is None:
        return empty_fig

    # Get the data from the .xlsx file
    elms, list_data = get_elements(folderpath_edx, with_plot=True)
    if element_edx not in elms:
        return empty_fig

    X_POS, Y_POS, ELM = make_heatmap(list_data, element_edx)

    if z_min is None:
        z_min = np.min(ELM)
    if z_max is None:
        z_max = np.max(ELM)


    fig = go.Figure(data=go.Heatmap(x=X_POS, y=Y_POS, z=ELM, colorscale="Plasma",
                                    colorbar=colorbar_layout(z_min, z_max, title=f'{element_edx} <br> at.%')
                                    ))

    if z_min is not None:
        fig.data[0].update(zmin=z_min)
    if z_max is not None:
        fig.data[0].update(zmax=z_max)

    return fig
